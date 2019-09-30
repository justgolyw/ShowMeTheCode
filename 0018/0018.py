#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@Author : yangwei.li
@Create date : 2019-09-30
@FileName : 0018.py
"""

"""
第 0018 题： 将第 0014 题中的 city.xlxs 文件中的内容写到 city.xml 文件中
"""
import openpyxl
from xml.dom import minidom
from collections import OrderedDict
import json


def xlxs2xml(xlxs_name):
    wb = openpyxl.load_workbook(xlxs_name)
    ws = wb.active
    dic = OrderedDict()

    for i in range(1,ws.max_row+1):
        key = int(ws.cell(i,1).value)
        value = str(ws.cell(i,2).value)
        dic[key] = value
    jsonObj = json.dumps(dic,ensure_ascii=False,indent=4)
    print(jsonObj)

    # 创建DOM树对象
    dom = minidom.Document()
    # 使用DOM对象来创建节点
    root = dom.createElement('root') # 根节点
    dom.appendChild(root)
    child = dom.createElement('citys')
    root.appendChild(child)
    comment = dom.createComment('城市信息')
    child.appendChild(comment)
    child_text = dom.createTextNode(jsonObj+"\n")
    child.appendChild(child_text)


    try:
        with open('city.xml','w',encoding='utf-8') as f:
            dom.writexml(f,indent='',addindent='\t',newl='\n',encoding='utf-8')
            print('写入xml成功')
    except Exception as e:
        print(e)





if __name__ == "__main__":
    xlxs2xml("city.xlsx")