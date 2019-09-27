#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
第 0014 题： 纯文本文件 student.txt为学生信息, 里面的内容（包括花括号）如下所示：
"""


import openpyxl
import json

# 本题考查json文件转化为excel文件
def json_to_excel(path):
    # 创建excel文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "example"
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            print(data)
            row_index, col_index = 1, 1
            for key, value in data.items():
                sheet.cell(row=row_index, column=1, value=key)
                sheet.cell(row=row_index, column=2, value=value)
                row_index += 1
            wb.save("city.xlsx")
            wb.close()
    except Exception as e:
        print(e)


if __name__ == "__main__":
    json_to_excel("city.txt")


