#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@Author : yangwei.li
@Create date : 2019-09-30
@FileName : 0019.py
"""
'''
第 0019 题： 将第 0016 题中的 numbers.xls 文件中的内容写到 numbers.xml 文件中
'''

import openpyxl
from xml.dom import minidom
from collections import OrderedDict
import json


def xlxs2xml(xlxs_name):
    wb = openpyxl.load_workbook(xlxs_name)
    ws = wb.active
    table = []
    table_t = []

    for i in range(1,ws.max_row+1):
        table = []
        for j in range(1,ws.max_column+1):
            value = int(ws.cell(i,j).value)
            table.append(value)
        table_t.append(table)

    jsonObj = json.dumps(table_t,ensure_ascii=False)
    print(jsonObj)

    # 创建DOM树对象
    dom = minidom.Document()
    # 使用DOM对象来创建节点
    root = dom.createElement('root') # 根节点
    dom.appendChild(root)
    child = dom.createElement('numbers')
    root.appendChild(child)
    comment = dom.createComment('数字信息')
    child.appendChild(comment)
    child_text = dom.createTextNode(jsonObj)
    child.appendChild(child_text)


    try:
        with open('number.xml','w',encoding='utf-8') as f:
            dom.writexml(f,indent='',addindent='\t',newl='\n',encoding='utf-8')
            print('写入xml成功')
    except Exception as e:
        print(e)


if __name__ == "__main__":
    xlxs2xml("number.xlsx")